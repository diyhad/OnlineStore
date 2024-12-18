import openpyxl
import sqlite3
import os
import re

CONN = sqlite3.connect("barrette.db")
CURSOR = CONN.cursor()


def extract_image_number(filename):
    match = re.search(r"/(\d+)\.(jpg|png)$", filename)
    if match:
        return int(match.group(1))
    else:
        return float("inf")


def get_images_dir(barcode, directories):
    for d in directories:
        if barcode in d:
            return d


def parse_table(folder_path):
    directories = os.listdir(folder_path)
    images_directories = []
    for d in directories:
        if ".xlsx" in d:
            continue
        barcodes = os.listdir(f"{folder_path}/{d}")
        for code in barcodes:
            images_directories.append(f"{folder_path}/{d}/{code}")

    table = [t for t in directories if ".xlsx" in t][0]
    wb = openpyxl.load_workbook(f"{folder_path}/{table}")
    sheet = wb.sheetnames
    if len(sheet) != 1:
        raise Exception("Table has more or less than 1 sheet")
    else:
        sheet = sheet[0]

    series_description = None

    for raw_row in wb[sheet].iter_rows(min_row=2, values_only=True):
        row = []
        for col in raw_row:
            if col and not str(col).startswith("="):
                if isinstance(col, str):
                    col = col.strip()
                row.append(col)
            else:
                row.append("")

        barcode, material_code, series, bg_name = row[0:4]
        series_description = row[4] if row[4] else series_description
        (
            product_type,
            en_name,
            short_description,
            description,
            usage,
            ingredients,
            benefits,
            size,
        ) = row[5:]

        product_folder = get_images_dir(str(barcode), images_directories)
        if product_folder:
            images_paths = [f"{product_folder}/{i}" for i in os.listdir(product_folder)]
            images_paths = sorted(images_paths, key=extract_image_number)

            image_paths = []
            for image in images_paths:
                if os.path.isfile(image):  # Ensure it's a file, not a directory
                    image_paths.append(image)
        else:
            image_paths = []

        CURSOR.execute(
            "INSERT INTO items (barcode, bg_name, en_name, material_code, "
            "series, series_description, product_type, short_description, "
            "description, usage, ingredients, benefits, size, images) VALUES"
            "(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ? ,? ,? ,?)",
            (
                barcode,
                bg_name,
                en_name,
                material_code,
                series,
                series_description,
                product_type,
                short_description,
                description,
                usage,
                ingredients,
                benefits,
                size,
                repr(image_paths),
            ),
        )


if __name__ == "__main__":

    CURSOR.execute(
        """
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            barcode REAL,
            bg_name TEXT NOT NULL,
            en_name TEXT,
            material_code TEXT,
            series TEXT,
            series_description TEXT,
            product_type TEXT,
            short_description TEXT,
            description TEXT,
            usage TEXT,
            ingredients TEXT,
            benefits TEXT,
            size TEXT,
            images TEXT
        )
    """
    )

    data_dir = "static/images"
    brands = os.listdir(data_dir)

    for brand in brands:
        path = f"{data_dir}/{brand}"
        folders = os.listdir(path) if os.path.isdir(path) else []

        for folder in folders:
            if "& Product info" in folder:
                parse_table(f"{path}/{folder}")

    CONN.commit()
    CONN.close()
